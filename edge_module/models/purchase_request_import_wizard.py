# -*- coding: utf-8 -*-
import base64
import io
import logging
from datetime import datetime

from odoo import api, fields, models, _
from odoo.exceptions import UserError, ValidationError

_logger = logging.getLogger(__name__)

try:
    import openpyxl
except ImportError:
    _logger.debug('Cannot `import openpyxl`. Please install it.')
    openpyxl = None

# Mapping from Excel Selection values to Odoo keys (Adjust based on exact Excel values)
DELIVERY_LOC_MAP = {
    'Edge Autonomy HSV': 'edge_slo',
    'Other': 'other',
}
RESALE_MAP = {
    'Resale': 'resale',
    'No Resale': 'no_resale',
}
PROD_STOPPAGE_MAP = { # Example, adjust based on actual excel values
    'Yes': True,
    'No': False,
    'TRUE': True,
    'FALSE': False,
     True: True,
     False: False,
}
PURCHASE_TYPE_MAP = {
    'Direct Material': 'direct_materials', # Check exact text in Excel
    'Direct Service': 'direct_services', # Check exact text in Excel
    'Indirect Material': 'indirect_materials', # Check exact text in Excel
    'Indirect Service': 'indirect_services', # Check exact text in Excel
    # Add variations if necessary (e.g., lowercase)
    'direct material': 'direct_materials',
    'direct service': 'direct_services',
    'indirect material': 'indirect_materials',
    'indirect service': 'indirect_services',
}
EXPENSE_TYPE_MAP = {
    # Map ALL possible Excel values from the 'Expense Type' column to Odoo keys
    'Subcontractors/Consultants/Outside Professionals': 'subcontractors',
    'Inventory (Raw Materials)': 'raw_materials',
    'Consumables': 'consumables',
    'Small Tooling': 'small_tooling',
    'Manufacturing Supplies': 'manufacturing_supplies',
    'Engineering Supplies': 'engineering_supplies',
    'Office Supplies': 'office_supplies',
    'Facilities - Building Supplies': 'building_supplies',
    'Facilities - Janitorial': 'janitorial',
    'Facilities - Phones/Internet/Communications': 'communications',
    'Facilities - Utilities & Waste': 'utilities',
    'Flight Ops Materials & Supplies': 'flight_ops',
    'IT Hardware': 'it_hardware',
    'IT Software': 'it_software',
    'IT Services': 'it_services',
    'Repairs & Maintenance': 'repairs',
    'Business Development Expenses': 'business_dev',
    'Conference/Seminar/Training Fees': 'training',
    'Licenses & Permits': 'licenses',
    'Vehicle Supplies': 'vehicle',
    'Equipment Rental': 'equipment_rental',
    'Employee Morale Costs': 'employee_morale',
    'Safety Supplies': 'safety',
    'Marketing Expenses': 'marketing',
    'Recruiting Costs': 'recruiting',
    'Shipping & Freight, Packaging Supplies': 'shipping',
    'Direct Award Materials (Cost of Good Sold)': 'direct_award',
    'Capital Expenditures, non-IR&D (>$2,500)': 'capex',
    # Add lowercase/alternative mappings if needed
    'inventory (raw materials)': 'raw_materials',
}


class PurchaseRequestImportWizard(models.TransientModel):
    _name = 'purchase.request.import.wizard'
    _description = 'Purchase Request Import Wizard'

    file_data = fields.Binary(string='Excel File', required=True)
    file_name = fields.Char(string='File Name')

    def _get_value_from_cell(self, sheet, cell_coord):
        """ Helper to get cell value safely """
        cell = sheet[cell_coord]
        return cell.value if cell else None

    def _find_employee_by_name(self, name):
        """ Find hr.employee """
        if not name:
            return False
        Employee = self.env['hr.employee']
        # Adjust search criteria if needed (e.g., search by related user's name)
        employee = Employee.search([('name', '=ilike', name)], limit=1)
        if not employee:
             _logger.warning(f"Employee not found for name: {name}")
             # Optional: raise UserError(_(f"Employee '{name}' not found."))
        return employee.id

    def _find_partner_by_name(self, name):
        """ Find res.partner (supplier) """
        if not name:
            return False
        Partner = self.env['res.partner']
        partner = Partner.search([
            ('name', '=ilike', name),
            ('supplier_rank', '>', 0)
        ], limit=1)
        if not partner:
            _logger.warning(f"Suggested Vendor (res.partner) not found for name: {name}")
            # Optional: raise UserError(_(f"Suggested Vendor '{name}' not found."))
        return partner.id

    def _find_product_by_code_or_name(self, identifier):
        """ Find product.product """
        if not identifier:
            return False
        Product = self.env['product.product']
        # Prioritize search by Internal Reference (Part Number), then by Name
        product = Product.search([
            ('default_code', '=ilike', identifier),
            ('purchase_ok', '=', True)
            ], limit=1)
        if not product:
            product = Product.search([
                ('name', '=ilike', identifier),
                ('purchase_ok', '=', True)
                ], limit=1)
        if not product:
             raise UserError(_(f"Product not found or not purchasable for identifier: '{identifier}'"))
        return product.id

    def _find_uom_by_name(self, name):
        """ Find uom.uom """
        if not name:
            return False # UOM is required on line
        Uom = self.env['uom.uom']
        uom = Uom.search([('name', '=ilike', name)], limit=1)
        if not uom:
            raise UserError(_(f"Unit of Measure (UOM) not found for name: '{name}'"))
        return uom.id

    def _find_job_by_name_or_code(self, identifier):
        """ Find job """
        if not identifier:
            return False # Job is required on line
        Job = self.env['job'] # Make sure 'job' is the correct model name
        # Adjust search criteria based on how jobs are identified (name, code, etc.)
        job = Job.search(['|', ('name', '=ilike', identifier), ('job_number', '=ilike', identifier)], limit=1) # Example search
        if not job:
            raise UserError(_(f"Job not found for identifier: '{identifier}'"))
        return job.id

    def _parse_date(self, date_value):
        """ Parse date, handling Excel's potential numeric dates """
        if not date_value:
            return False
        if isinstance(date_value, datetime):
            return date_value.date()
        if isinstance(date_value, (int, float)):
            # Assuming standard Excel date origin (may need adjustment)
            try:
                 # openpyxl might handle this conversion automatically depending on settings
                 # If not, implement Excel serial date conversion logic here
                 # This is a simplified example, robust Excel date conversion can be complex
                 from datetime import datetime, timedelta
                 # Excel base date varies (1900 or 1904)
                 base_date = datetime(1899, 12, 30) # Common base for Windows Excel
                 return (base_date + timedelta(days=date_value)).date()
            except Exception as e:
                 _logger.error(f"Could not convert Excel numeric date {date_value}: {e}")
                 return False
        if isinstance(date_value, str):
             try:
                 # Attempt common string formats
                 return datetime.strptime(date_value, '%Y-%m-%d').date()
             except ValueError:
                 try:
                      return datetime.strptime(date_value, '%m/%d/%Y').date()
                 except ValueError:
                    _logger.warning(f"Could not parse date string: {date_value}")
                    return False
        return False


    def action_import_file(self):
        self.ensure_one()
        if not openpyxl:
            raise UserError(_("The 'openpyxl' library is not installed. Please install it."))
        if not self.file_data:
            raise UserError(_("Please upload an Excel file."))

        try:
            workbook = openpyxl.load_workbook(filename=io.BytesIO(base64.b64decode(self.file_data)), data_only=True)
            sheet = workbook.active # Or specify sheet name: workbook['Sheet1']

            # --- Extract Header Data ---
            pr_vals = {}
            notes_list = [] # Collect notes

            # B12: Suggested Vendor
            vendor_name = self._get_value_from_cell(sheet, 'B12')
            if vendor_name:
                pr_vals['partner_id'] = self._find_partner_by_name(vendor_name)

            # B13: Supplier Contact Info
            supplier_contact = self._get_value_from_cell(sheet, 'B13')
            if supplier_contact:
                notes_list.append(f"Supplier Contact Info: {supplier_contact}")

            # B14: Production Stoppage
            prod_stoppage_val = self._get_value_from_cell(sheet, 'B14')
            if prod_stoppage_val is not None: # Check explicitly for None in case value is False
                 pr_vals['production_stoppage'] = PROD_STOPPAGE_MAP.get(prod_stoppage_val, False) # Default to False if not found/empty


            # B15: Need Date
            need_date_val = self._get_value_from_cell(sheet, 'B15')
            parsed_need_date = self._parse_date(need_date_val)
            if not parsed_need_date:
                raise UserError(_("Need by Date (Cell B15) is missing or invalid format."))
            pr_vals['need_by_date'] = parsed_need_date

            # B16/B17/B18/B19: Delivery Info
            delivery_loc_val = self._get_value_from_cell(sheet, 'B16')
            delivery_address_mapped = DELIVERY_LOC_MAP.get(delivery_loc_val)
            if not delivery_address_mapped:
                 _logger.warning(f"Delivery Location '{delivery_loc_val}' in cell B16 is not recognized. Defaulting or skipping.")
                 # Decide default or raise error: raise UserError(_(f"Delivery Location '{delivery_loc_val}' (Cell B16) is invalid."))
                 pr_vals['deliver_to_address'] = 'edge_slo' # Example default
            else:
                pr_vals['deliver_to_address'] = delivery_address_mapped

            delivery_poc_name = self._get_value_from_cell(sheet, 'B18')
            delivery_address_other = self._get_value_from_cell(sheet, 'B17')
            delivery_phone_other = self._get_value_from_cell(sheet, 'B19')

            if pr_vals.get('deliver_to_address') == 'other':
                 pr_vals['deliver_to_other'] = delivery_poc_name
                 pr_vals['deliver_to_other_address'] = delivery_address_other
                 pr_vals['deliver_to_other_phone'] = delivery_phone_other
                 if not delivery_poc_name or not delivery_address_other:
                      _logger.warning("Delivery location is 'Other' but External Recipient Name or Address is missing (B18, B17).")
                      # Optional: raise UserError
            else:
                 # If Edge SLO, try to find internal recipient
                 pr_vals['deliver_to'] = self._find_employee_by_name(delivery_poc_name)


            # B20: Invoice Approver
            invoice_approver = self._get_value_from_cell(sheet, 'B20')
            if invoice_approver:
                notes_list.append(f"Invoice Approver: {invoice_approver}")

            # B21: Resale Designation
            resale_val = self._get_value_from_cell(sheet, 'B21')
            resale_mapped = RESALE_MAP.get(resale_val)
            if not resale_mapped:
                 raise UserError(_(f"Resale Designation '{resale_val}' (Cell B21) is missing or invalid."))
            pr_vals['resale_designation'] = resale_mapped

            # B22: Notes/Links
            notes_links = self._get_value_from_cell(sheet, 'B22')
            if notes_links:
                notes_list.append(str(notes_links)) # Ensure it's a string

            pr_vals['requester_notes'] = "\n".join(notes_list)

            # --- Set Defaults / Contextual Data ---
            pr_vals['requester_id'] = self.env.user.id
            # Try to find Originator (hr.employee) linked to the current user
            originator_employee = self.env['hr.employee'].search([('user_id', '=', self.env.user.id)], limit=1)
            if not originator_employee:
                # Fallback or error - Originator is required!
                # Option 1: Raise Error
                 raise UserError(_("Could not find an Employee record linked to your user account. Cannot set the required 'Originator' field."))
                # Option 2: Set a default if possible (e.g., a specific admin/placeholder, but risky)
                # pr_vals['originator'] = self.env.ref('your_module.default_originator_employee').id # Example
            else:
                pr_vals['originator'] = originator_employee.id


            # --- Extract Line Items ---
            request_line_ids = []
            # Find the header row (assuming 'Purchase Type' is a unique enough header)
            header_row_num = 0
            for row_idx in range(1, sheet.max_row + 1):
                 # Check multiple columns for robustness
                 cell_a = sheet.cell(row=row_idx, column=1).value # Col A: Purchase Type
                 cell_b = sheet.cell(row=row_idx, column=2).value # Col B: Edge Internal Part Number
                 if cell_a and isinstance(cell_a, str) and "purchase type" in cell_a.lower() and \
                    cell_b and isinstance(cell_b, str) and "part number" in cell_b.lower():
                     header_row_num = row_idx
                     break

            if header_row_num == 0:
                raise UserError(_("Could not find the header row for line items (looking for 'Purchase Type', 'Edge Internal Part Number'). Check the template."))

            # Map header names to column indexes (letters to numbers: A=1, B=2, ...)
            col_map = {}
            for col_idx in range(1, sheet.max_column + 1):
                 header_val = sheet.cell(row=header_row_num, column=col_idx).value
                 if header_val and isinstance(header_val, str):
                      col_map[header_val.strip().lower()] = col_idx # Store lowercase header


            # Define expected columns based on mapping (use lowercase)
            expected_cols = {
                'purchase type': 'purchase_type',
                'edge internal part number': 'product_id',
                'description': 'name',
                'uom': 'product_uom_id',
                'qty': 'quantity',
                'est unit price': 'price_unit',
                'job': 'job',
                'expense type': 'expense_type',
                'pop start': 'pop_start',
                'pop end': 'pop_end',
                'mfr pn': 'manufacturer_number',
                'mfr': 'manufacturer',
                'cage code': 'cage_code',
                # Add 'drawing revision' if needed
            }

            # Check if all required columns exist in the header
            missing_cols = [hdr for hdr in expected_cols if hdr not in col_map]
            if missing_cols:
                 raise UserError(_(f"Missing required columns in the Excel header: {', '.join(missing_cols)}"))


            # Start reading from the row after the header
            for row_idx in range(header_row_num + 1, sheet.max_row + 1):
                line_vals = {}
                # Check if row seems empty (e.g., first few required cells are blank)
                part_num_col_idx = col_map.get('edge internal part number')
                qty_col_idx = col_map.get('qty')
                part_num_val = sheet.cell(row=row_idx, column=part_num_col_idx).value if part_num_col_idx else None
                qty_val = sheet.cell(row=row_idx, column=qty_col_idx).value if qty_col_idx else None

                if not part_num_val and not qty_val: # Basic check for empty row
                     _logger.info(f"Skipping potentially empty row {row_idx}.")
                     continue # Skip empty rows

                # Extract data based on mapped columns
                for header, field_name in expected_cols.items():
                    col_idx = col_map.get(header)
                    if col_idx:
                         cell_val = sheet.cell(row=row_idx, column=col_idx).value
                         # Clean/Convert data
                         if cell_val is not None: # Only process non-empty cells
                             if field_name == 'product_id':
                                 line_vals[field_name] = self._find_product_by_code_or_name(str(cell_val).strip())
                             elif field_name == 'product_uom_id':
                                 line_vals[field_name] = self._find_uom_by_name(str(cell_val).strip())
                             elif field_name == 'job':
                                  line_vals[field_name] = self._find_job_by_name_or_code(str(cell_val).strip())
                             elif field_name == 'quantity':
                                 try:
                                     line_vals[field_name] = float(cell_val)
                                     if line_vals[field_name] <= 0:
                                          raise UserError(_(f"Row {row_idx}: Quantity must be positive ('{cell_val}')."))
                                 except (ValueError, TypeError):
                                     raise UserError(_(f"Row {row_idx}: Invalid Quantity value '{cell_val}'."))
                             elif field_name == 'price_unit':
                                 try:
                                     line_vals[field_name] = float(cell_val)
                                 except (ValueError, TypeError):
                                     raise UserError(_(f"Row {row_idx}: Invalid Estimated Unit Cost value '{cell_val}'."))
                             elif field_name == 'purchase_type':
                                 mapped_pt = PURCHASE_TYPE_MAP.get(str(cell_val).strip())
                                 if not mapped_pt:
                                      # Try lowercase match as fallback
                                      mapped_pt = PURCHASE_TYPE_MAP.get(str(cell_val).strip().lower())
                                      if not mapped_pt:
                                        raise UserError(_(f"Row {row_idx}: Invalid or unmapped Purchase Type '{cell_val}'. Check mapping and template value."))
                                 line_vals[field_name] = mapped_pt
                             elif field_name == 'expense_type':
                                 mapped_et = EXPENSE_TYPE_MAP.get(str(cell_val).strip())
                                 if not mapped_et:
                                     # Try lowercase match as fallback
                                     mapped_et = EXPENSE_TYPE_MAP.get(str(cell_val).strip().lower())
                                     if not mapped_et:
                                         raise UserError(_(f"Row {row_idx}: Invalid or unmapped Expense Type '{cell_val}'. Check mapping and template value."))
                                 line_vals[field_name] = mapped_et
                             elif field_name in ['pop_start', 'pop_end']:
                                 parsed_date = self._parse_date(cell_val)
                                 if parsed_date: # POP dates might be optional
                                     line_vals[field_name] = parsed_date
                                 elif cell_val: # Log if there was a value we couldn't parse
                                      _logger.warning(f"Row {row_idx}: Could not parse date for {field_name}: {cell_val}")
                             else: # For text fields like name, manufacturer, etc.
                                 line_vals[field_name] = str(cell_val).strip()
                         elif field_name in ['name', 'quantity', 'price_unit', 'product_uom_id', 'job', 'purchase_type', 'expense_type', 'product_id']:
                              # Check required fields that ended up empty
                              if field_name != 'product_id' or part_num_val: # Don't raise if part num was also blank (likely end of data)
                                   raise UserError(_(f"Row {row_idx}: Missing required value for column '{header}'."))


                # Add description if product found but description column was empty
                if 'product_id' in line_vals and not line_vals.get('name'):
                    product = self.env['product.product'].browse(line_vals['product_id'])
                    line_vals['name'] = product.name_get()[0][1] # Get product display name

                # Add UOM if product found and UOM column was empty/invalid
                if 'product_id' in line_vals and not line_vals.get('product_uom_id'):
                    product = self.env['product.product'].browse(line_vals['product_id'])
                    line_vals['product_uom_id'] = product.uom_po_id.id or product.uom_id.id
                    if not line_vals['product_uom_id']:
                         raise UserError(_(f"Row {row_idx}: Could not determine default Unit of Measure for product '{product.display_name}'. Please specify UOM in the file."))


                # Validate required fields for the line before adding
                required_line_fields = ['product_id', 'name', 'quantity', 'price_unit', 'product_uom_id', 'job', 'purchase_type', 'expense_type']
                missing_req = [f for f in required_line_fields if not line_vals.get(f)]
                if missing_req:
                     # If part number was originally blank, assume end of relevant data
                     if not part_num_val:
                         _logger.info(f"Stopping line processing at row {row_idx} due to missing part number and required fields: {', '.join(missing_req)}.")
                         break # Stop processing lines
                     else:
                         raise UserError(_(f"Row {row_idx}: Missing required line values for: {', '.join(missing_req)}"))


                if line_vals: # Only add if we extracted some data
                    request_line_ids.append((0, 0, line_vals))

            if not request_line_ids:
                 raise UserError(_("No valid line items found in the Excel file."))

            pr_vals['request_line_ids'] = request_line_ids

            # --- Create Purchase Request ---
            purchase_request = self.env['purchase.request'].create(pr_vals)
            _logger.info(f"Successfully created Purchase Request {purchase_request.name} from file {self.file_name}")

            # --- Return action to view the created PR ---
            return {
                'name': _('Created Purchase Request'),
                'type': 'ir.actions.act_window',
                'res_model': 'purchase.request',
                'res_id': purchase_request.id,
                'view_mode': 'form',
                'target': 'current', # Or 'main' to open in main content area
            }

        except FileNotFoundError:
            raise UserError(_("Error: Could not find the file."))
        except openpyxl.utils.exceptions.InvalidFileException:
             raise UserError(_("Invalid file format. Please upload a valid Excel (.xlsx) file."))
        except UserError as e:
            raise e # Re-raise UserErrors directly
        except ValidationError as e:
             raise UserError(_("Validation Error: %s") % str(e.args[0]))
        except Exception as e:
            _logger.exception("Error processing Purchase Request import file.")
            raise UserError(_("An unexpected error occurred during import: %s") % str(e))